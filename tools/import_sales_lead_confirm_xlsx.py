#!/usr/bin/env python3
"""
将「销售线索确认」Excel 导入表 sales_lead_confirm。

默认列（首行表头，可与 Excel 一致）：
  customer_name, sales, lead_row_id, sales_from_row_id, register_date_latest, register_date_sales

用法（项目根）:
  python tools/import_sales_lead_confirm_xlsx.py
  python tools/import_sales_lead_confirm_xlsx.py "E:/path/销售线索确认.xlsx"
  python tools/import_sales_lead_confirm_xlsx.py --dry-run

导入前需已建表：python tools/apply_dispatch_schema.py
依赖：openpyxl（pip install openpyxl）
"""
from __future__ import annotations

import argparse
import configparser
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "销售线索确认.xlsx"


def _cell_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def main() -> int:
    ap = argparse.ArgumentParser(description="导入销售线索确认 Excel → sales_lead_confirm")
    ap.add_argument(
        "xlsx",
        nargs="?",
        default=str(DEFAULT_XLSX),
        help=f"xlsx 路径，默认 {DEFAULT_XLSX}",
    )
    ap.add_argument("--dry-run", action="store_true", help="只解析打印行数，不写库")
    args = ap.parse_args()

    path = Path(args.xlsx)
    if not path.is_file():
        print(f"文件不存在: {path}", file=sys.stderr)
        return 1

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("请先 pip install openpyxl", file=sys.stderr)
        return 1

    cfg_path = ROOT / "config.ini"
    cp = configparser.ConfigParser()
    cp.read(cfg_path, encoding="utf-8")
    if not cp.has_section("database"):
        print("config.ini 缺少 [database]", file=sys.stderr)
        return 1

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        print("Excel 无数据")
        return 1

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {name.lower().replace(" ", "_"): i for i, name in enumerate(header)}
    # 兼容中英文列名
    def col(*names: str) -> int | None:
        for n in names:
            key = n.lower().replace(" ", "_")
            if key in idx:
                return idx[key]
        return None

    i_cn = col("customer_name", "客户名称", "客户")
    i_sales = col("sales", "售前")
    i_lid = col("lead_row_id")
    i_sid = col("sales_from_row_id")
    i_rdl = col("register_date_latest")
    i_rds = col("register_date_sales")
    if i_cn is None:
        print(f"未找到 customer_name 列，表头={header}", file=sys.stderr)
        return 1

    batch: list[tuple] = []
    for row in rows[1:]:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        cn = row[i_cn] if i_cn < len(row) else None
        customer_name = str(cn).strip() if cn is not None else ""
        if not customer_name:
            continue
        sales_v = row[i_sales] if i_sales is not None and i_sales < len(row) else None
        sales = None if sales_v is None else str(sales_v).strip()
        if sales == "":
            sales = None

        def gint(ii: int | None) -> int | None:
            if ii is None or ii >= len(row):
                return None
            v = row[ii]
            if v is None or v == "":
                return None
            try:
                return int(v)
            except Exception:
                return None

        batch.append(
            (
                customer_name,
                sales,
                gint(i_lid),
                gint(i_sid),
                _cell_date(row[i_rdl] if i_rdl is not None and i_rdl < len(row) else None),
                _cell_date(row[i_rds] if i_rds is not None and i_rds < len(row) else None),
                1,
            )
        )

    print(f"[parse] 有效数据行: {len(batch)}（列 customer_name 非空）")
    if args.dry_run:
        print("[dry-run] 不写库")
        return 0

    import pymysql

    db = cp["database"]
    conn = pymysql.connect(
        host=db["host"],
        port=int(db["port"]),
        user=db["user"],
        password=db["password"],
        database=db["database"],
        charset=db.get("charset", "utf8mb4"),
    )
    try:
        with conn.cursor() as cur:
            cur.execute("TRUNCATE TABLE `sales_lead_confirm`")
            if batch:
                cur.executemany(
                    "INSERT INTO `sales_lead_confirm` "
                    "(`customer_name`,`sales`,`lead_row_id`,`sales_from_row_id`,"
                    "`register_date_latest`,`register_date_sales`,`enabled`) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    batch,
                )
        conn.commit()
        print(f"[db] TRUNCATE + INSERT {len(batch)} 行 → sales_lead_confirm")
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
